from sqlalchemy.orm import Session
from typing import Tuple, List, Optional
from datetime import datetime
import os
import openpyxl
from app.models.models import Debtor, Batch, BatchStatus, DebtorStatus


class ImportService:
    """Excel import service for debtors"""
    
    @staticmethod
    def validate_excel_file(file_path: str) -> Tuple[bool, str, int]:
        """
        Validate Excel file.
        Returns (is_valid, error_message, row_count)
        """
        if not os.path.exists(file_path):
            return False, "File not found", 0
        
        if not file_path.endswith(('.xlsx', '.xls')):
            return False, "File must be an Excel file (.xlsx or .xls)", 0
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            row_count = ws.max_row - 1  # Exclude header
            wb.close()
            
            if row_count <= 0:
                return False, "Excel file is empty", 0
            
            return True, "", row_count
        except Exception as e:
            return False, f"Error reading Excel file: {str(e)}", 0
    
    @staticmethod
    def validate_row(row: dict, row_num: int) -> Tuple[bool, str]:
        """
        Validate a single row of debtor data.
        Expected columns: name, id_card, phone, address (optional), debt_amount (optional), remark (optional)
        """
        required_fields = ['name', 'id_card', 'phone']
        
        for field in required_fields:
            if field not in row or not row[field]:
                return False, f"Row {row_num}: Missing required field '{field}'"
        
        # Validate ID card length
        id_card = str(row.get('id_card', ''))
        if len(id_card) < 15 or len(id_card) > 18:
            return False, f"Row {row_num}: Invalid ID card length"
        
        # Validate phone length
        phone = str(row.get('phone', ''))
        if len(phone) < 11 or len(phone) > 20:
            return False, f"Row {row_num}: Invalid phone length"
        
        # Validate debt_amount if provided
        if 'debt_amount' in row and row['debt_amount']:
            try:
                amount = float(row['debt_amount'])
                if amount < 0:
                    return False, f"Row {row_num}: Debt amount cannot be negative"
            except (ValueError, TypeError):
                return False, f"Row {row_num}: Invalid debt amount"
        
        return True, ""
    
    @staticmethod
    def import_debtors(db: Session, batch_id: int, file_path: str, 
                      created_by: int = None) -> Tuple[int, int, List[str]]:
        """
        Import debtors from Excel file into a batch.
        Returns (success_count, fail_count, errors)
        """
        # Verify batch exists
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return 0, 0, ["Batch not found"]
        
        if batch.status != BatchStatus.PENDING:
            return 0, 0, [f"Batch is not in pending status (current: {batch.status.value})"]
        
        # Validate file
        is_valid, error, row_count = ImportService.validate_excel_file(file_path)
        if not is_valid:
            return 0, 0, [error]
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            success_count = 0
            fail_count = 0
            errors = []
            duplicate_id_cards = set()
            
            # Update batch status to processing
            batch.status = BatchStatus.PROCESSING
            db.commit()
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = dict(zip(headers, row))
                
                # Validate row
                is_valid, error = ImportService.validate_row(row_dict, row_num)
                if not is_valid:
                    fail_count += 1
                    errors.append(error)
                    continue
                
                # Check for duplicate in import
                id_card = str(row_dict['id_card'])
                if id_card in duplicate_id_cards:
                    fail_count += 1
                    errors.append(f"Row {row_num}: Duplicate ID card {id_card} in file")
                    continue
                
                # Check for duplicate in database
                existing = db.query(Debtor).filter(Debtor.id_card == id_card).first()
                if existing:
                    fail_count += 1
                    errors.append(f"Row {row_num}: ID card {id_card} already exists in system")
                    continue
                
                try:
                    debtor = Debtor(
                        name=str(row_dict['name']),
                        id_card=id_card,
                        phone=str(row_dict['phone']),
                        address=str(row_dict.get('address', '')) if row_dict.get('address') else None,
                        debt_amount=float(row_dict['debt_amount']) if row_dict.get('debt_amount') else 0.0,
                        remark=str(row_dict.get('remark', '')) if row_dict.get('remark') else None,
                        batch_id=batch_id,
                        status=DebtorStatus.ACTIVE
                    )
                    db.add(debtor)
                    duplicate_id_cards.add(id_card)
                    success_count += 1
                except Exception as e:
                    fail_count += 1
                    errors.append(f"Row {row_num}: Failed to create debtor - {str(e)}")
            
            wb.close()
            
            # Update batch
            batch.total_count = success_count + fail_count
            batch.success_count = success_count
            batch.fail_count = fail_count
            batch.status = BatchStatus.COMPLETED if success_count > 0 else BatchStatus.FAILED
            db.commit()
            
            return success_count, fail_count, errors
            
        except Exception as e:
            batch.status = BatchStatus.FAILED
            db.commit()
            return success_count, fail_count, errors + [f"Import failed: {str(e)}"]
