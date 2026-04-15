import openpyxl
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from io import BytesIO


class ExcelImporter:
    """
    Excel file importer for debtor data.
    Supports .xlsx and .xls formats.
    """
    
    REQUIRED_HEADERS = [
        'debtor_number',
        'name',
    ]
    
    OPTIONAL_HEADERS = [
        'id_card',
        'phone',
        'email',
        'bank_name',
        'bank_account',
        'bank_account_name',
        'address',
        'remark',
        'status',
        'overdue_amount',
        'overdue_days',
    ]
    
    ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS
    
    STATUS_MAPPING = {
        'active': 'active',
        'overdue': 'overdue',
        'cleared': 'cleared',
        'legal': 'legal',
        '正常': 'active',
        '逾期': 'overdue',
        '已结清': 'cleared',
        '法务': 'legal',
    }
    
    def __init__(self, file_path: Optional[str] = None, file_content: Optional[bytes] = None):
        """
        Initialize importer with either file path or file content bytes.
        """
        self.file_path = file_path
        self.file_content = file_content
        self.workbook = None
        self.sheet = None
        self.headers: List[str] = []
        self.errors: List[str] = []
        self.warnings: List[str] = []
    
    def load(self) -> bool:
        """
        Load the Excel file.
        Returns True if successful, False otherwise.
        """
        try:
            if self.file_content:
                self.workbook = openpyxl.load_workbook(BytesIO(self.file_content))
            elif self.file_path:
                self.workbook = openpyxl.load_workbook(self.file_path)
            else:
                self.errors.append("No file path or content provided")
                return False
            
            self.sheet = self.workbook.active
            self._parse_headers()
            return True
        except Exception as e:
            self.errors.append(f"Failed to load file: {str(e)}")
            return False
    
    def _parse_headers(self):
        """
        Parse headers from the first row.
        """
        try:
            header_row = list(self.sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            self.headers = [str(h).strip().lower() if h else '' for h in header_row]
            
            missing_required = set(self.REQUIRED_HEADERS) - set(self.headers)
            if missing_required:
                self.errors.append(f"Missing required headers: {', '.join(missing_required)}")
        except Exception as e:
            self.errors.append(f"Failed to parse headers: {str(e)}")
    
    def validate_headers(self) -> bool:
        """
        Validate that required headers are present.
        """
        missing_required = set(self.REQUIRED_HEADERS) - set(self.headers)
        if missing_required:
            self.errors.append(f"Missing required headers: {', '.join(missing_required)}")
            return False
        return True
    
    def _get_column_index(self, header: str) -> Optional[int]:
        """
        Get the column index for a given header name.
        """
        try:
            return self.headers.index(header.lower())
        except ValueError:
            return None
    
    def _get_cell_value(self, row: List, header: str, default: Any = None) -> Any:
        """
        Get cell value for a given header from a row.
        """
        col_idx = self._get_column_index(header)
        if col_idx is None or col_idx >= len(row):
            return default
        value = row[col_idx]
        if value is None:
            return default
        if isinstance(value, datetime):
            return value
        return str(value).strip() if value else default
    
    def _validate_status(self, status: str) -> Optional[str]:
        """
        Validate and normalize status value.
        """
        if not status:
            return 'active'
        status_lower = status.lower().strip()
        return self.STATUS_MAPPING.get(status_lower, 'active')
    
    def _validate_phone(self, phone: str) -> Optional[str]:
        """
        Validate phone number format.
        """
        if not phone:
            return None
        phone = str(phone).strip()
        phone = ''.join(c for c in phone if c.isdigit() or c == '+')
        if len(phone) < 7 or len(phone) > 15:
            return None
        return phone
    
    def _validate_amount(self, value: Any) -> int:
        """
        Validate and convert amount to integer.
        """
        if not value or value == '':
            return 0
        try:
            if isinstance(value, (int, float)):
                return int(value)
            return int(float(str(value)))
        except (ValueError, TypeError):
            return 0
    
    def parse_rows(self) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Parse all data rows from the Excel file.
        
        Returns:
            Tuple of (valid_rows, invalid_rows)
            valid_rows: List of validated debtor dictionaries
            invalid_rows: List of row data with error messages
        """
        if not self.validate_headers():
            return [], []
        
        valid_rows = []
        invalid_rows = []
        
        total_rows = self.sheet.max_row - 1
        
        for row_idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            debtor_number = self._get_cell_value(row, 'debtor_number')
            name = self._get_cell_value(row, 'name')
            
            if not debtor_number:
                invalid_rows.append({
                    'row': row_idx,
                    'data': row,
                    'errors': ['Missing debtor_number']
                })
                continue
            
            if not name:
                invalid_rows.append({
                    'row': row_idx,
                    'data': row,
                    'errors': ['Missing name']
                })
                continue
            
            phone = self._get_cell_value(row, 'phone')
            if phone:
                phone = self._validate_phone(phone)
            
            status = self._get_cell_value(row, 'status', 'active')
            
            debtor_data = {
                'debtor_number': debtor_number,
                'name': name,
                'id_card': self._get_cell_value(row, 'id_card'),
                'phone': phone,
                'email': self._get_cell_value(row, 'email'),
                'bank_name': self._get_cell_value(row, 'bank_name'),
                'bank_account': self._get_cell_value(row, 'bank_account'),
                'bank_account_name': self._get_cell_value(row, 'bank_account_name'),
                'address': self._get_cell_value(row, 'address'),
                'remark': self._get_cell_value(row, 'remark'),
                'status': self._validate_status(status),
                'overdue_amount': self._validate_amount(self._get_cell_value(row, 'overdue_amount')),
                'overdue_days': self._validate_amount(self._get_cell_value(row, 'overdue_days')),
                'row_number': row_idx,
            }
            
            valid_rows.append(debtor_data)
        
        return valid_rows, invalid_rows
    
    def get_summary(self) -> Dict[str, Any]:
        """
        Get import summary.
        """
        return {
            'total_rows': self.sheet.max_row - 1 if self.sheet else 0,
            'headers': self.headers,
            'errors': self.errors,
            'warnings': self.warnings,
        }


class ExcelExporter:
    """
    Excel file exporter for debtor data.
    """
    
    DEFAULT_HEADERS = [
        'ID',
        '债务人编号',
        '姓名',
        '身份证',
        '手机号',
        '邮箱',
        '开户行',
        '银行账号',
        '账户名',
        '地址',
        '状态',
        '逾期金额',
        '逾期天数',
        '查询次数',
        '最后查询时间',
        '创建时间',
    ]
    
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "债务人数据"
        self.row_idx = 1
    
    def write_header(self, headers: Optional[List[str]] = None):
        """
        Write header row.
        """
        headers = headers or self.DEFAULT_HEADERS
        for col_idx, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=self.row_idx, column=col_idx)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
        self.row_idx += 1
    
    def write_row(self, data: Dict[str, Any]):
        """
        Write a data row.
        """
        row_data = [
            data.get('id', ''),
            data.get('debtor_number', ''),
            data.get('name', ''),
            data.get('id_card', ''),
            data.get('phone', ''),
            data.get('email', ''),
            data.get('bank_name', ''),
            data.get('bank_account', ''),
            data.get('bank_account_name', ''),
            data.get('address', ''),
            data.get('status', ''),
            data.get('overdue_amount', 0),
            data.get('overdue_days', 0),
            data.get('query_count', 0),
            data.get('last_query_at', ''),
            data.get('created_at', ''),
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            self.sheet.cell(row=self.row_idx, column=col_idx, value=value)
        
        self.row_idx += 1
    
    def write_rows(self, data_list: List[Dict[str, Any]]):
        """
        Write multiple data rows.
        """
        for data in data_list:
            self.write_row(data)
    
    def save(self, file_path: str):
        """
        Save the workbook to a file.
        """
        self.workbook.save(file_path)
    
    def get_bytes(self) -> bytes:
        """
        Get the workbook as bytes.
        """
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
